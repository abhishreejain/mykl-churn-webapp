"""Build cleaned yearly wide datasets from raw files."""

from __future__ import annotations

import calendar
import csv
import logging
from collections import defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from churn_model.cleaning import clean_mobile, standardize_raw_wide, standardize_name, standardize_state
from churn_model.config import ProjectConfig
from churn_model.io import discover_raw_files, infer_year_from_path, read_csv, read_excel_sheets, write_csv
from churn_model.schema import MONTH_COLUMNS, apply_column_aliases, build_alias_lookup, infer_layout, validate_raw_wide_schema

LOGGER = logging.getLogger(__name__)


def build_yearly_datasets_from_config(cfg: ProjectConfig) -> list[Path]:
    """Build cleaned year-level wide datasets and quality reports."""
    return build_yearly_datasets(
        raw_dir=cfg.path("paths.raw_data_dir"),
        output_dir=cfg.path("paths.processed_data_dir"),
        reports_dir=cfg.path("paths.reports_dir"),
        alias_config=cfg.get("schema.column_aliases", {}),
        blank_tokens=cfg.get("cleaning.blank_tokens", []),
        mobile_config=cfg.get("cleaning.mobile", {}),
        duplicate_config=cfg.get("cleaning.duplicates", {}),
        monthly_dump_config=cfg.get("cleaning.monthly_dump", {}),
    )


def build_yearly_datasets(
    raw_dir: Path,
    output_dir: Path,
    reports_dir: Path,
    alias_config: dict[str, list[str]] | None = None,
    blank_tokens: list[str] | None = None,
    mobile_config: dict[str, Any] | None = None,
    duplicate_config: dict[str, Any] | None = None,
    monthly_dump_config: dict[str, list[str]] | None = None,
) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    reports_dir.mkdir(parents=True, exist_ok=True)
    mobile_config = mobile_config or {}
    duplicate_config = duplicate_config or {}

    cleaned_by_year: dict[int, list[pd.DataFrame]] = defaultdict(list)
    source_rows: list[dict[str, object]] = []
    cleaning_rows: list[dict[str, object]] = []
    duplicate_reports: list[pd.DataFrame] = []
    outputs: list[Path] = []

    for file_path in discover_raw_files(raw_dir):
        file_year = infer_year_from_path(file_path)
        source_display_path = _display_path(file_path)
        if file_path.suffix.lower() == ".xlsx":
            stream_result = process_xlsx_wide_streaming(
                file_path=file_path,
                source_display_path=source_display_path,
                inferred_year=file_year,
                output_dir=output_dir,
                alias_config=alias_config,
                blank_tokens=blank_tokens or [],
                mobile_config=mobile_config,
                duplicate_config=duplicate_config,
                monthly_dump_config=monthly_dump_config or {},
            )
            source_rows.extend(stream_result["source_rows"])
            cleaning_rows.extend(stream_result["cleaning_rows"])
            duplicate_reports.extend(stream_result["duplicate_reports"])
            outputs.extend(stream_result["outputs"])
            continue

        sheets = {"csv": read_csv(file_path)} if file_path.suffix.lower() == ".csv" else read_excel_sheets(file_path)
        for sheet_name, raw_df in sheets.items():
            aliased = apply_column_aliases(raw_df, alias_config)
            layout = infer_layout([str(col) for col in aliased.columns])
            source_rows.append(
                {
                    "file_path": source_display_path,
                    "sheet_name": sheet_name,
                    "inferred_year": file_year,
                    "row_count": len(aliased),
                    "column_names": "|".join(map(str, aliased.columns)),
                    "detected_layout": layout,
                }
            )
            if layout == "Jan-Dec wide":
                if file_year is None:
                    raise ValueError(f"Cannot clean Jan-Dec wide source without inferred year: {file_path}")
                validate_raw_wide_schema(aliased, f"{file_path}::{sheet_name}")
                cleaned, report = standardize_raw_wide(
                    aliased,
                    year=file_year,
                    source_file=source_display_path,
                    sheet_name=sheet_name,
                    blank_tokens=blank_tokens,
                    prefer_last_n_digits=int(mobile_config.get("prefer_last_n_digits", 10)),
                    valid_mobile_lengths=tuple(mobile_config.get("valid_lengths", [10])),
                )
                cleaned_by_year[file_year].append(cleaned)
                cleaning_rows.append(report)
            elif layout == "monthly-long candidate":
                assembled = assemble_monthly_dump(
                    aliased,
                    source_file=source_display_path,
                    sheet_name=sheet_name,
                    inferred_year=file_year,
                    monthly_dump_config=monthly_dump_config or {},
                )
                for year, wide in assembled.items():
                    validate_raw_wide_schema(wide, f"{file_path}::{sheet_name} assembled {year}")
                    cleaned, report = standardize_raw_wide(
                        wide,
                        year=year,
                        source_file=source_display_path,
                        sheet_name=sheet_name,
                        blank_tokens=blank_tokens,
                        prefer_last_n_digits=int(mobile_config.get("prefer_last_n_digits", 10)),
                        valid_mobile_lengths=tuple(mobile_config.get("valid_lengths", [10])),
                    )
                    cleaned_by_year[year].append(cleaned)
                    cleaning_rows.append(report)
            else:
                LOGGER.warning("Skipping source with unsupported layout: %s :: %s", file_path, sheet_name)

    if not cleaned_by_year:
        if outputs:
            write_csv(pd.DataFrame(source_rows), reports_dir / "raw_source_layout_report.csv")
            write_csv(pd.DataFrame(cleaning_rows), reports_dir / "cleaning_summary.csv")
            all_duplicates = (
                pd.concat(duplicate_reports, ignore_index=True) if duplicate_reports else pd.DataFrame()
            )
            write_csv(all_duplicates, reports_dir / "duplicate_diagnostics.csv")
            write_cleaning_assumptions_report(
                reports_dir / "cleaning_assumptions.md",
                duplicate_config=duplicate_config,
                mobile_config=mobile_config,
                blank_tokens=blank_tokens or [],
            )
            return outputs
        raise FileNotFoundError(f"No readable raw sources found under {raw_dir}")

    for year, frames in sorted(cleaned_by_year.items()):
        year_df = pd.concat(frames, ignore_index=True)
        resolved, duplicate_report = resolve_duplicates(year_df, duplicate_config)
        duplicate_reports.append(duplicate_report)
        output = output_dir / f"yearly_{year}_cleaned.csv"
        write_csv(resolved, output)
        outputs.append(output)

    write_csv(pd.DataFrame(source_rows), reports_dir / "raw_source_layout_report.csv")
    write_csv(pd.DataFrame(cleaning_rows), reports_dir / "cleaning_summary.csv")
    all_duplicates = pd.concat(duplicate_reports, ignore_index=True) if duplicate_reports else pd.DataFrame()
    write_csv(all_duplicates, reports_dir / "duplicate_diagnostics.csv")
    write_cleaning_assumptions_report(
        reports_dir / "cleaning_assumptions.md",
        duplicate_config=duplicate_config,
        mobile_config=mobile_config,
        blank_tokens=blank_tokens or [],
    )
    return outputs


def process_xlsx_wide_streaming(
    file_path: Path,
    source_display_path: str,
    inferred_year: int | None,
    output_dir: Path,
    alias_config: dict[str, list[str]] | None,
    blank_tokens: list[str],
    mobile_config: dict[str, Any],
    duplicate_config: dict[str, Any],
    monthly_dump_config: dict[str, list[str]],
) -> dict[str, list[Any]]:
    """Fast path for large XLSX Jan-Dec wide sheets."""
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    outputs: list[Path] = []
    source_rows: list[dict[str, object]] = []
    cleaning_rows: list[dict[str, object]] = []
    duplicate_reports: list[pd.DataFrame] = []
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            try:
                raw_headers = next(rows)
            except StopIteration:
                continue
            headers = normalize_headers(raw_headers, alias_config)
            layout = infer_layout(headers)
            source_rows.append(
                {
                    "file_path": source_display_path,
                    "sheet_name": sheet_name,
                    "inferred_year": inferred_year,
                    "row_count": max(sheet.max_row - 1, 0),
                    "column_names": "|".join(headers),
                    "detected_layout": layout,
                }
            )
            if layout != "Jan-Dec wide":
                # Fall back to pandas for uncommon monthly-long XLSX sources.
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                aliased = apply_column_aliases(df, alias_config)
                assembled = assemble_monthly_dump(
                    aliased,
                    source_file=source_display_path,
                    sheet_name=sheet_name,
                    inferred_year=inferred_year,
                    monthly_dump_config=monthly_dump_config,
                )
                for year, wide in assembled.items():
                    cleaned, report = standardize_raw_wide(
                        wide,
                        year,
                        source_display_path,
                        sheet_name,
                        blank_tokens,
                        int(mobile_config.get("prefer_last_n_digits", 10)),
                        tuple(mobile_config.get("valid_lengths", [10])),
                    )
                    resolved, dup = resolve_duplicates(cleaned, duplicate_config)
                    output = output_dir / f"yearly_{year}_cleaned.csv"
                    write_csv(resolved, output)
                    outputs.append(output)
                    cleaning_rows.append(report)
                    duplicate_reports.append(dup)
                continue
            if inferred_year is None:
                raise ValueError(f"Cannot clean Jan-Dec wide source without inferred year: {file_path}")
            cleaned_rows, report, duplicate_report = clean_and_resolve_stream(
                rows=rows,
                headers=headers,
                year=inferred_year,
                source_file=source_display_path,
                sheet_name=sheet_name,
                blank_tokens=blank_tokens,
                mobile_config=mobile_config,
                duplicate_config=duplicate_config,
            )
            output = output_dir / f"yearly_{inferred_year}_cleaned.csv"
            write_cleaned_rows_csv(cleaned_rows, output)
            outputs.append(output)
            cleaning_rows.append(report)
            duplicate_reports.append(duplicate_report)
    finally:
        workbook.close()
    return {
        "outputs": outputs,
        "source_rows": source_rows,
        "cleaning_rows": cleaning_rows,
        "duplicate_reports": duplicate_reports,
    }


def clean_and_resolve_stream(
    rows: Any,
    headers: list[str],
    year: int,
    source_file: str,
    sheet_name: str,
    blank_tokens: list[str],
    mobile_config: dict[str, Any],
    duplicate_config: dict[str, Any],
) -> tuple[list[dict[str, object]], dict[str, object], pd.DataFrame]:
    blank_set = {str(token).strip().lower() for token in blank_tokens}
    prefer_last = int(mobile_config.get("prefer_last_n_digits", 10))
    valid_lengths = tuple(mobile_config.get("valid_lengths", [10]))
    groups: dict[tuple[int, str], dict[str, object]] = {}
    invalid_rows: list[dict[str, object]] = []
    report: dict[str, object] = {
        "source_file": source_file,
        "source_sheet": sheet_name,
        "source_year": year,
        "input_rows": 0,
    }
    for column in ["State", "Name", "MobileNo1", *MONTH_COLUMNS]:
        report[f"{column}_blank_or_null_count"] = 0
        if column in MONTH_COLUMNS:
            report[f"{column}_numeric_coerce_to_null_count"] = 0
    trimmed_count = 0
    invalid_count = 0

    index = {header: position for position, header in enumerate(headers)}
    required = ["State", "Name", "MobileNo1", *MONTH_COLUMNS]
    missing = [column for column in required if column not in index]
    if missing:
        raise ValueError(f"{source_file}::{sheet_name} missing required columns: {missing}")

    for source_row_number, values in enumerate(rows, start=2):
        report["input_rows"] += 1
        raw = {column: values[pos] if pos < len(values) else None for column, pos in index.items()}
        state = _blank_to_none(raw.get("State"), blank_set)
        name = _blank_to_none(raw.get("Name"), blank_set)
        mobile_raw = _blank_to_none(raw.get("MobileNo1"), blank_set)
        if state is None:
            report["State_blank_or_null_count"] += 1
        if name is None:
            report["Name_blank_or_null_count"] += 1
        if mobile_raw is None:
            report["MobileNo1_blank_or_null_count"] += 1
        mobile = clean_mobile(mobile_raw, prefer_last, valid_lengths)
        if mobile.issue.startswith("trimmed"):
            trimmed_count += 1
        if not mobile.is_valid:
            invalid_count += 1

        scan_values: dict[str, float | None] = {}
        for month in MONTH_COLUMNS:
            value = _blank_to_none(raw.get(month), blank_set)
            if value is None:
                report[f"{month}_blank_or_null_count"] += 1
                scan_values[month] = None
                continue
            try:
                scan_values[month] = float(value)
            except (TypeError, ValueError):
                report[f"{month}_numeric_coerce_to_null_count"] += 1
                scan_values[month] = None

        base_row = {
            "State": standardize_state(state),
            "Name": standardize_name(name),
            "MobileNo1": mobile_raw,
            "mobile_digits": mobile.digits,
            "mobile_id": mobile.mobile_id,
            "mobile_is_valid": mobile.is_valid,
            "mobile_cleaning_issue": mobile.issue,
            "source_year": year,
            "source_file": source_file,
            "source_sheet": sheet_name,
            "source_row_number": source_row_number,
            **scan_values,
        }
        if not mobile.is_valid or not mobile.mobile_id:
            base_row["duplicate_resolution_rule"] = "invalid_mobile_kept_each_source_row_flagged"
            base_row["source_duplicate_row_count"] = 1
            invalid_rows.append(base_row)
            continue
        key = (year, mobile.mobile_id)
        if key not in groups:
            groups[key] = {
                **base_row,
                "_state_values": [],
                "_name_values": [],
                "_raw_mobile_values": [],
                "_source_files": set(),
                "_source_sheets": set(),
                "_duplicate_row_count": 0,
            }
            for month in MONTH_COLUMNS:
                groups[key][month] = 0.0
        group = groups[key]
        group["_duplicate_row_count"] = int(group["_duplicate_row_count"]) + 1
        group["_source_files"].add(source_file)
        group["_source_sheets"].add(sheet_name)
        if state is not None:
            group["_state_values"].append(standardize_state(state))
        if name is not None:
            group["_name_values"].append(standardize_name(name))
        if mobile_raw is not None:
            group["_raw_mobile_values"].append(mobile_raw)
        group["source_row_number"] = min(int(group["source_row_number"]), source_row_number)
        for month in MONTH_COLUMNS:
            if scan_values[month] is not None:
                group[month] = float(group[month]) + float(scan_values[month])

    report["invalid_mobile_count"] = invalid_count
    report["trimmed_mobile_count"] = trimmed_count
    resolved: list[dict[str, object]] = []
    duplicate_rows: list[dict[str, object]] = []
    for (source_year, mobile_id), group in groups.items():
        duplicate_count = int(group["_duplicate_row_count"])
        if duplicate_count > 1:
            duplicate_rows.append(
                {
                    "source_year": source_year,
                    "mobile_id": mobile_id,
                    "duplicate_row_count": duplicate_count,
                    "source_files": "|".join(sorted(group["_source_files"])),
                    "source_sheets": "|".join(sorted(group["_source_sheets"])),
                    "raw_mobile_values": "|".join(sorted({str(v) for v in group["_raw_mobile_values"]})),
                }
            )
        row = {key: value for key, value in group.items() if not str(key).startswith("_")}
        row["State"] = _representative_from_values(group["_state_values"])
        row["Name"] = _representative_from_values(group["_name_values"])
        row["MobileNo1"] = _representative_from_values(group["_raw_mobile_values"])
        row["source_file"] = "|".join(sorted(group["_source_files"]))
        row["source_sheet"] = "|".join(sorted(group["_source_sheets"]))
        row["duplicate_resolution_rule"] = duplicate_config.get("resolution_rule", "aggregate_by_mobile_year")
        row["source_duplicate_row_count"] = duplicate_count
        resolved.append(row)
    resolved.extend(invalid_rows)
    return resolved, report, pd.DataFrame(duplicate_rows)


def normalize_headers(raw_headers: tuple[object, ...], alias_config: dict[str, list[str]] | None) -> list[str]:
    lookup = build_alias_lookup(alias_config)
    headers: list[str] = []
    seen: set[str] = set()
    for header in raw_headers:
        stripped = str(header).strip() if header is not None else ""
        canonical = lookup.get("".join(ch for ch in stripped.lower() if ch.isalnum()), stripped)
        if canonical in seen:
            canonical = stripped
        seen.add(canonical)
        headers.append(canonical)
    return headers


def write_cleaned_rows_csv(rows: list[dict[str, object]], output: Path) -> None:
    columns = [
        "State",
        "Name",
        "MobileNo1",
        "mobile_digits",
        "mobile_id",
        "mobile_is_valid",
        "mobile_cleaning_issue",
        "source_year",
        "source_file",
        "source_sheet",
        "source_row_number",
        "duplicate_resolution_rule",
        "source_duplicate_row_count",
        *MONTH_COLUMNS,
    ]
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    LOGGER.info("Wrote %s rows to %s", len(rows), output)


def assemble_monthly_dump(
    df: pd.DataFrame,
    source_file: str,
    sheet_name: str,
    inferred_year: int | None,
    monthly_dump_config: dict[str, list[str]],
) -> dict[int, pd.DataFrame]:
    month_column = _find_first_present(df, monthly_dump_config.get("month_aliases", []))
    scan_column = _find_first_present(df, monthly_dump_config.get("scan_value_aliases", []))
    year_column = _find_first_present(df, monthly_dump_config.get("year_aliases", []))
    required = ["State", "Name", "MobileNo1"]
    missing = [col for col in required if col not in df.columns]
    if missing or month_column is None or scan_column is None:
        raise ValueError(
            f"Monthly dump source lacks required identity/month/scan columns: {source_file}::{sheet_name}"
        )
    work = df.copy()
    work["source_year"] = work[year_column].astype("Int64") if year_column else inferred_year
    if work["source_year"].isna().any():
        raise ValueError(f"Monthly dump source has rows without year and no year in path: {source_file}")
    work["month_name"] = work[month_column].map(_month_to_abbr)
    work["scan_value"] = pd.to_numeric(work[scan_column], errors="coerce")
    grouped = (
        work.groupby(["source_year", "State", "Name", "MobileNo1", "month_name"], dropna=False)["scan_value"]
        .sum(min_count=1)
        .reset_index()
    )
    wide = grouped.pivot_table(
        index=["source_year", "State", "Name", "MobileNo1"],
        columns="month_name",
        values="scan_value",
        aggfunc="sum",
        dropna=False,
    ).reset_index()
    result: dict[int, pd.DataFrame] = {}
    for year, year_df in wide.groupby("source_year", dropna=False):
        if pd.isna(year):
            continue
        output = year_df.drop(columns=["source_year"]).copy()
        for month in MONTH_COLUMNS:
            if month not in output.columns:
                output[month] = pd.NA
        result[int(year)] = output[["State", "Name", "MobileNo1", *MONTH_COLUMNS]]
    return result


def resolve_duplicates(
    cleaned: pd.DataFrame,
    duplicate_config: dict[str, Any],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    rule = duplicate_config.get("resolution_rule", "aggregate_by_mobile_year")
    if rule != "aggregate_by_mobile_year":
        raise ValueError(f"Unsupported duplicate resolution rule: {rule}")

    valid = cleaned[cleaned["mobile_is_valid"].fillna(False)].copy()
    invalid = cleaned[~cleaned["mobile_is_valid"].fillna(False)].copy()
    group_cols = duplicate_config.get("group_by", ["source_year", "mobile_id"])
    scan_aggregation = duplicate_config.get("scan_aggregation", "sum")
    if scan_aggregation != "sum":
        raise ValueError(f"Unsupported scan aggregation: {scan_aggregation}")

    duplicate_report = (
        valid.groupby(group_cols, dropna=False)
        .agg(
            duplicate_row_count=("source_row_number", "size"),
            source_files=("source_file", lambda s: "|".join(sorted(set(map(str, s))))),
            source_sheets=("source_sheet", lambda s: "|".join(sorted(set(map(str, s))))),
            raw_mobile_values=("MobileNo1", lambda s: "|".join(sorted(set(map(str, s.dropna()))))),
        )
        .reset_index()
    )
    duplicate_report = duplicate_report[duplicate_report["duplicate_row_count"] > 1].copy()

    if valid.empty:
        resolved_valid = valid
    else:
        aggregations: dict[str, object] = {
            "State": _representative_value,
            "Name": _representative_value,
            "MobileNo1": _representative_value,
            "mobile_digits": _representative_value,
            "mobile_is_valid": "first",
            "mobile_cleaning_issue": _representative_value,
            "source_file": lambda s: "|".join(sorted(set(map(str, s)))),
            "source_sheet": lambda s: "|".join(sorted(set(map(str, s)))),
            "source_row_number": "min",
        }
        for month in MONTH_COLUMNS:
            aggregations[month] = "sum"
        resolved_valid = valid.groupby(group_cols, dropna=False, as_index=False).agg(aggregations)
        resolved_valid["duplicate_resolution_rule"] = rule
        resolved_valid["source_duplicate_row_count"] = valid.groupby(group_cols, dropna=False).size().to_numpy()

    if not invalid.empty:
        invalid = invalid.copy()
        invalid["duplicate_resolution_rule"] = "invalid_mobile_kept_each_source_row_flagged"
        invalid["source_duplicate_row_count"] = 1

    resolved = pd.concat([resolved_valid, invalid], ignore_index=True, sort=False)
    ordered = [
        "State",
        "Name",
        "MobileNo1",
        "mobile_digits",
        "mobile_id",
        "mobile_is_valid",
        "mobile_cleaning_issue",
        "source_year",
        "source_file",
        "source_sheet",
        "source_row_number",
        "duplicate_resolution_rule",
        "source_duplicate_row_count",
        *MONTH_COLUMNS,
    ]
    return resolved[[col for col in ordered if col in resolved.columns]], duplicate_report


def write_cleaning_assumptions_report(
    path: Path,
    duplicate_config: dict[str, Any],
    mobile_config: dict[str, Any],
    blank_tokens: list[str],
) -> None:
    content = f"""# Cleaning Assumptions

## Column Aliases

Raw columns are normalized through configured aliases in `configs/base.yaml`. The audited default mapping is `State`, `Name`, `MobileNo1`, and `Jan` through `Dec`.

## Mobile Cleaning

- Trim whitespace.
- Remove all non-digit characters.
- Retain `mobile_digits` and canonical `mobile_id`.
- Prefer a valid {mobile_config.get("prefer_last_n_digits", 10)}-digit canonical representation when the raw digit string is longer.
- Flag invalid or unusable mobiles in `mobile_is_valid` and `mobile_cleaning_issue`.
- Do not use raw mobile or cleaned mobile as model features.

## Blank and Numeric Cleaning

- Configured blank tokens: {blank_tokens}
- Blank/null counts and numeric coercion counts are written to `reports/cleaning_summary.csv`.
- Monthly scan columns are converted to numeric values after reporting configured blank/null tokens.

## Duplicate Resolution

- Configured rule: `{duplicate_config.get("resolution_rule", "aggregate_by_mobile_year")}`.
- Grouping: `{duplicate_config.get("group_by", ["source_year", "mobile_id"])}`.
- Monthly scan aggregation: `{duplicate_config.get("scan_aggregation", "sum")}`.
- Identity resolution: `{duplicate_config.get("identity_resolution", "most_frequent_nonblank_then_first")}`.
- Invalid mobile rows are kept as separate source rows and flagged rather than merged into a customer identifier.
- Duplicate diagnostics are written to `reports/duplicate_diagnostics.csv`.

This rule is a data-engineering rule for repeated cleaned identifiers. It does not define churn, activity, decline, risk, or potential logic.
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def safe_name(value: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in value).strip("_").lower() or "sheet"


def load_yearly_files(processed_dir: Path) -> pd.DataFrame:
    files = sorted(processed_dir.glob("yearly_*_cleaned.csv"))
    if not files:
        files = sorted(processed_dir.glob("yearly_*.csv"))
    if not files:
        raise FileNotFoundError(f"No yearly cleaned files found in {processed_dir}")
    return pd.concat((pd.read_csv(path) for path in files), ignore_index=True)


def _find_first_present(df: pd.DataFrame, candidates: list[str]) -> str | None:
    normalized = {str(col).strip().lower(): str(col) for col in df.columns}
    for candidate in candidates:
        found = normalized.get(str(candidate).strip().lower())
        if found is not None:
            return found
    return None


def _month_to_abbr(value: object) -> str | pd.NA:
    if pd.isna(value):
        return pd.NA
    if isinstance(value, pd.Timestamp):
        return calendar.month_abbr[value.month]
    text = str(value).strip()
    if not text:
        return pd.NA
    if text.isdigit():
        month_number = int(text)
        if 1 <= month_number <= 12:
            return calendar.month_abbr[month_number]
    parsed = pd.to_datetime(text, errors="coerce")
    if not pd.isna(parsed):
        return calendar.month_abbr[int(parsed.month)]
    lowered = text[:3].title()
    return lowered if lowered in MONTH_COLUMNS else pd.NA


def _representative_value(series: pd.Series) -> object:
    nonblank = series.dropna()
    nonblank = nonblank[nonblank.astype(str).str.strip() != ""]
    if nonblank.empty:
        return pd.NA
    modes = nonblank.mode(dropna=True)
    return modes.iloc[0] if not modes.empty else nonblank.iloc[0]


def _blank_to_none(value: object, blank_set: set[str]) -> object | None:
    if pd.isna(value):
        return None
    text = str(value).strip()
    return None if text.lower() in blank_set else value


def _representative_from_values(values: list[object]) -> object:
    cleaned = [value for value in values if value is not None and str(value).strip() != ""]
    if not cleaned:
        return None
    counts: dict[str, int] = {}
    originals: dict[str, object] = {}
    for value in cleaned:
        key = str(value)
        counts[key] = counts.get(key, 0) + 1
        originals.setdefault(key, value)
    best = sorted(counts.items(), key=lambda item: (-item[1], item[0]))[0][0]
    return originals[best]


def _display_path(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(Path.cwd().resolve()))
    except ValueError:
        return str(path)
